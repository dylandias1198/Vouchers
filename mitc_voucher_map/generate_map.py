#!/usr/bin/env python3
"""Generate data.js for MITC voucher India map + Users Demo from voucher_new.xlsx."""

import json
import os
from collections import defaultdict
from openpyxl import load_workbook

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "voucher_new.xlsx")
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "data.js")
LIFETIME_HIST_PATH = os.path.join(os.path.dirname(__file__), "lifetime-histograms.js")

ALL_STATE_CODES = [
    "in-an", "in-ap", "in-ar", "in-as", "in-br", "in-cg", "in-ch", "in-dd",
    "in-dl", "in-dn", "in-ga", "in-gj", "in-hp", "in-hr", "in-jh", "in-jk",
    "in-ka", "in-kl", "in-ld", "in-mh", "in-ml", "in-mn", "in-mp", "in-mz",
    "in-nl", "in-od", "in-pb", "in-py", "in-rj", "in-sk", "in-tn", "in-tr",
    "in-ts", "in-uk", "in-up", "in-wb",
]

PIN_PREFIX_3 = {
    **{p: "in-dl" for p in range(110, 111)},
    **{p: "in-hr" for p in list(range(121, 130)) + list(range(131, 137))},
    **{p: "in-pb" for p in range(140, 161) if p not in (140, 160)},
    140: "in-ch",
    160: "in-ch",
    **{p: "in-hp" for p in range(171, 178)},
    **{p: "in-jk" for p in range(180, 195)},
    194: "in-jk",
    **{p: "in-uk" for p in range(244, 264) if p < 263},
    263: "in-uk",
    **{p: "in-up" for p in range(201, 244)},
    **{p: "in-up" for p in range(264, 286)},
    **{p: "in-rj" for p in range(301, 346)},
    **{p: "in-gj" for p in range(360, 396)},
    396: "in-dn",
    362: "in-dn",
    395: "in-dn",
    **{p: "in-mh" for p in range(400, 445)},
    **{p: "in-mp" for p in range(450, 489)},
    **{p: "in-cg" for p in range(490, 500)},
    **{p: "in-ts" for p in range(500, 510)},
    **{p: "in-ap" for p in range(510, 536)},
    **{p: "in-ka" for p in range(560, 591)},
    **{p: "in-tn" for p in range(600, 644)},
    **{p: "in-py" for p in range(605, 608)},
    **{p: "in-kl" for p in range(670, 696)},
    682: "in-ld",
    **{p: "in-wb" for p in range(700, 744)},
    744: "in-an",
    **{p: "in-od" for p in range(751, 770)},
    **{p: "in-as" for p in range(781, 789)},
    790: "in-ml",
    791: "in-ml",
    792: "in-ml",
    793: "in-tr",
    794: "in-tr",
    795: "in-mz",
    796: "in-mz",
    797: "in-mn",
    798: "in-nl",
    799: "in-tr",
    737: "in-sk",
    **{p: "in-br" for p in range(800, 856)},
    **{p: "in-jh" for p in range(813, 835) if p >= 813},
    **{p: "in-jh" for p in range(825, 835)},
    403: "in-ga",
    786: "in-ar",
    787: "in-ar",
    788: "in-ar",
    789: "in-ar",
}

PIN_PREFIX_2 = {
    11: "in-dl", 12: "in-hr", 13: "in-hr", 14: "in-pb", 15: "in-pb", 16: "in-pb",
    17: "in-hp", 18: "in-jk", 19: "in-jk", 20: "in-up", 21: "in-up", 22: "in-up",
    23: "in-up", 24: "in-uk", 25: "in-up", 26: "in-up", 27: "in-up", 28: "in-up",
    30: "in-rj", 31: "in-rj", 32: "in-rj", 33: "in-rj", 34: "in-rj",
    36: "in-gj", 37: "in-gj", 38: "in-gj", 39: "in-gj",
    40: "in-mh", 41: "in-mh", 42: "in-mh", 43: "in-mh", 44: "in-mh",
    45: "in-mp", 46: "in-mp", 47: "in-mp", 48: "in-mp", 49: "in-cg",
    50: "in-ts", 51: "in-ap", 52: "in-ap", 53: "in-ap",
    56: "in-ka", 57: "in-ka", 58: "in-ka", 59: "in-ka",
    60: "in-tn", 61: "in-tn", 62: "in-tn", 63: "in-tn", 64: "in-tn",
    67: "in-kl", 68: "in-kl", 69: "in-kl",
    70: "in-wb", 71: "in-wb", 72: "in-wb", 73: "in-wb", 74: "in-wb",
    75: "in-od", 76: "in-od", 77: "in-od", 78: "in-as", 79: "in-as",
    80: "in-br", 81: "in-br", 82: "in-br", 83: "in-jh", 84: "in-jh", 85: "in-br",
}


def normalize_pin(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s.lower() in ("nan", "none"):
        return None
    if "." in s:
        s = s.split(".")[0]
    digits = "".join(c for c in s if c.isdigit())
    if not digits:
        return None
    if len(digits) > 6:
        digits = digits[:6]
    return digits.zfill(6)


def pincode_to_state_code(pin):
    pin = normalize_pin(pin)
    if not pin or len(pin) != 6:
        return None
    p3 = int(pin[:3])
    if p3 in PIN_PREFIX_3:
        return PIN_PREFIX_3[p3]
    return PIN_PREFIX_2.get(int(pin[:2]))


GENDER_CANONICAL = {
    "male": "Male",
    "m": "Male",
    "female": "Female",
    "f": "Female",
    "others": "Others",
    "other": "Others",
    "unknown": "Unknown",
}


def normalize_gender(g):
    if g is None:
        return "Unknown"
    s = str(g).strip()
    if not s or s.lower() in ("nan", "none", ""):
        return "Unknown"
    return GENDER_CANONICAL.get(s.lower(), s.title())


def standard_gender_order(genders):
    preferred = ["Female", "Male", "Others", "Unknown"]
    seen = set()
    ordered = []
    for g in preferred:
        if g in genders and g not in seen:
            ordered.append(g)
            seen.add(g)
    for g in sorted(genders):
        if g not in seen:
            ordered.append(g)
            seen.add(g)
    return ordered


def normalize_age(a):
    if a is None:
        return "Under 18 or Unknown"
    s = str(a).strip()
    return s if s else "Under 18 or Unknown"


def normalize_txn_date(txn_val, timestamp_txn=None):
    if txn_val is not None:
        if hasattr(txn_val, "strftime"):
            return txn_val.strftime("%Y-%m-%d")
        s = str(txn_val).strip()
        if s and s.lower() not in ("nan", "none", ""):
            return s[:10]
    if timestamp_txn is not None:
        s = str(timestamp_txn).strip()
        if s and s.lower() not in ("nan", "none", ""):
            return s[:10] if len(s) >= 10 else s[:7]
    return None


def normalize_mitc_flag(flag):
    if flag is None:
        return None
    s = str(flag).strip().upper()
    if s == "MITC":
        return "MITC"
    if s in ("NON-MITC", "NON MITC", "NONMITC"):
        return "NON-MITC"
    return None


def load_excel_rows():
    path = os.path.abspath(EXCEL_PATH)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}
    zip_key = next((h for h in headers if h and "zip" in str(h).lower()), "current_zipCode")
    mitc_rows = []
    user_records = []
    lifetime_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        mitc = normalize_mitc_flag(row[col["mitc_flag"]])
        if mitc is None:
            continue
        state = pincode_to_state_code(row[col[zip_key]])
        age = normalize_age(row[col.get("age_category", -1)] if "age_category" in col else None)
        gender = normalize_gender(row[col.get("gender", -1)] if "gender" in col else None)
        uuid = row[col.get("um_uuid", -1)] if "um_uuid" in col else None
        if uuid is None or str(uuid).strip() == "":
            continue
        uuid = str(uuid).strip()
        state_key = state if state else "unknown"
        user_records.append({"m": mitc, "a": age, "g": gender, "s": state_key, "u": uuid})
        amount = row[col.get("voucherAmount")]
        try:
            amount = float(amount) if amount is not None else 0.0
        except (TypeError, ValueError):
            amount = 0.0
        lifetime_rows.append({
            "mitc": mitc,
            "age": age,
            "gender": gender,
            "uuid": uuid,
            "amount": amount,
        })
        if mitc == "MITC":
            txn_date = normalize_txn_date(
                row[col["txn"]] if "txn" in col else None,
                row[col.get("timestamp_txn")] if "timestamp_txn" in col else None,
            )
            mitc_rows.append({
                "zip": row[col[zip_key]],
                "age": age,
                "gender": gender,
                "date": txn_date,
                "uuid": uuid,
                "amount": amount,
            })
    wb.close()
    return mitc_rows, user_records, lifetime_rows


def build_voucher_aggregates(rows):
    counts = defaultdict(lambda: defaultdict(int))
    unmapped = 0
    for r in rows:
        state = pincode_to_state_code(r["zip"])
        if not state:
            unmapped += 1
            continue
        age = normalize_age(r["age"])
        gender = normalize_gender(r["gender"])
        for a in ("All", age):
            for g in ("All", gender):
                counts[f"{a}|{g}"][state] += 1
    return counts, unmapped


def build_trend_aggregates(rows):
    counts = defaultdict(lambda: defaultdict(int))
    all_dates = set()
    skipped = 0
    for r in rows:
        d = r.get("date")
        if not d:
            skipped += 1
            continue
        all_dates.add(d)
        age = normalize_age(r["age"])
        gender = normalize_gender(r["gender"])
        for a in ("All", age):
            for g in ("All", gender):
                counts[f"{a}|{g}"][d] += 1
    trend_dates = sorted(all_dates)
    trend_data = {
        key: [lookup.get(d, 0) for d in trend_dates]
        for key, lookup in sorted(counts.items())
    }
    return trend_data, trend_dates, skipped


LIFETIME_BIN_EDGES = [0, 500, 1000, 2000, 5000, 10000, 20000, 50000, 100000, 250000, float("inf")]


def lifetime_bin_label(lo, hi):
    if hi == float("inf"):
        return f"{lo:,.0f}+"
    return f"{lo:,.0f}–{hi:,.0f}"


def bin_lifetime_counts(users, edges):
    n_bins = len(edges) - 1
    counts = [0] * n_bins
    for u in users:
        t = u["t"]
        for i in range(n_bins):
            lo = edges[i]
            hi = edges[i + 1]
            if hi == float("inf") or hi is None:
                if t >= lo:
                    counts[i] += 1
                    break
            elif lo <= t < hi:
                counts[i] += 1
                break
    return counts


def build_lifetime_data(rows):
    user_totals = defaultdict(float)
    user_meta = {}
    for r in rows:
        uid = r.get("uuid")
        if not uid:
            continue
        user_totals[uid] += r.get("amount", 0) or 0
        user_meta[uid] = {"a": r["age"], "g": r["gender"], "m": r["mitc"]}

    lifetime_users = [
        {
            "t": round(user_totals[uid], 2),
            "a": user_meta[uid]["a"],
            "g": user_meta[uid]["g"],
            "m": user_meta[uid]["m"],
        }
        for uid in user_totals
    ]
    labels = [
        lifetime_bin_label(LIFETIME_BIN_EDGES[i], LIFETIME_BIN_EDGES[i + 1])
        for i in range(len(LIFETIME_BIN_EDGES) - 1)
    ]
    return lifetime_users, labels


def build_lifetime_histograms(users, ages, genders):
    mitc_values = ["MITC", "NON-MITC", "All"]
    age_values = ["All"] + list(ages)
    gender_values = ["All"] + list(genders)
    n_bins = len(LIFETIME_BIN_EDGES) - 1
    histograms = {}
    user_totals = {}
    for mitc in mitc_values:
        for age in age_values:
            for gender in gender_values:
                filtered = [
                    u
                    for u in users
                    if (mitc == "All" or u["m"] == mitc)
                    and (age == "All" or u["a"] == age)
                    and (gender == "All" or u["g"] == gender)
                ]
                key = f"{mitc}|{age}|{gender}"
                histograms[key] = bin_lifetime_counts(filtered, LIFETIME_BIN_EDGES)
                user_totals[key] = len(filtered)
    return histograms, user_totals


def to_voucher_data(counts):
    value_by_key = {}
    for key, state_counts in counts.items():
        value_by_key[key] = {sc: v for sc, v in state_counts.items()}
    result = {}
    for key, lookup in sorted(value_by_key.items()):
        result[key] = [
            {"state_code": sc, "value": lookup.get(sc, 0)}
            for sc in ALL_STATE_CODES
        ]
    return result


def main():
    print("Loading", EXCEL_PATH)
    mitc_rows, user_records, lifetime_rows = load_excel_rows()
    print(f"MITC voucher rows: {len(mitc_rows):,}")
    print(f"User demo rows (MITC + NON-MITC, with state + uuid): {len(user_records):,}")

    counts, unmapped = build_voucher_aggregates(mitc_rows)
    all_key = "All|All"
    total_mapped = sum(counts[all_key].values())
    states_covered = len([v for v in counts[all_key].values() if v > 0])
    ages = sorted({normalize_age(r["age"]) for r in mitc_rows})
    genders = standard_gender_order({normalize_gender(r["gender"]) for r in mitc_rows})
    demo_ages = sorted({r["a"] for r in user_records})
    demo_genders = standard_gender_order({r["g"] for r in user_records})
    voucher_data = to_voucher_data(counts)
    trend_data, trend_dates, trend_skipped = build_trend_aggregates(mitc_rows)
    trend_records = [
        {"d": r["date"], "a": r["age"], "g": r["gender"]}
        for r in mitc_rows
        if r.get("date")
    ]
    lifetime_users, lifetime_labels = build_lifetime_data(lifetime_rows)
    lifetime_histograms, lifetime_user_totals = build_lifetime_histograms(
        lifetime_users, demo_ages, demo_genders
    )
    lifetime_mitc_n = sum(1 for u in lifetime_users if u["m"] == "MITC")
    lifetime_non_n = sum(1 for u in lifetime_users if u["m"] == "NON-MITC")

    print(f"Mapped vouchers: {total_mapped:,} ({unmapped:,} without valid pincode)")
    print(f"States with voucher data: {states_covered}")
    print(f"Trend dates: {len(trend_dates)} ({trend_skipped:,} MITC rows without txn date)")
    print(f"Lifetime users — MITC: {lifetime_mitc_n:,}, NON-MITC: {lifetime_non_n:,}, all: {len(lifetime_users):,}")

    js = (
        "// Generated from voucher_new.xlsx\n"
        "const allStateCodes = " + json.dumps(ALL_STATE_CODES) + ";\n"
        "const voucherData = " + json.dumps(voucher_data, indent=2) + ";\n"
        "const userRecords = " + json.dumps(user_records) + ";\n"
        "const trendDates = " + json.dumps(trend_dates) + ";\n"
        "const trendData = " + json.dumps(trend_data) + ";\n"
        "const trendRecords = " + json.dumps(trend_records) + ";\n"
        f"const statesCovered = {states_covered};\n"
        f"const ageCategories = {json.dumps(ages)};\n"
        f"const genders = {json.dumps(genders)};\n"
        f"const demoAgeCategories = {json.dumps(demo_ages)};\n"
        f"const demoGenders = {json.dumps(demo_genders)};\n"
        'const mitcFlagOptions = ["MITC", "NON-MITC"];\n'
    )
    lifetime_js = (
        "// Generated from voucher_new.xlsx — load before data.js\n"
        "const lifetimeBinLabels = " + json.dumps(lifetime_labels) + ";\n"
        "const lifetimeFilterAges = " + json.dumps(demo_ages) + ";\n"
        "const lifetimeFilterGenders = " + json.dumps(demo_genders) + ";\n"  # canonical: Female, Male, Others, Unknown
        "const lifetimeHistograms = " + json.dumps(lifetime_histograms) + ";\n"
        "const lifetimeUserTotals = " + json.dumps(lifetime_user_totals) + ";\n"
    )
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write(js)
    with open(LIFETIME_HIST_PATH, "w", encoding="utf-8") as f:
        f.write(lifetime_js)
    print("Wrote", OUTPUT_PATH)
    print("Wrote", LIFETIME_HIST_PATH)


if __name__ == "__main__":
    main()
